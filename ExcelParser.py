import openpyxl
import datetime
import CalendarEvent



class ExcelParser():

    def __init__(self, path):
        
        wb = openpyxl.load_workbook(path) 
        self.sheet = wb[wb.sheetnames[0]]

        self.max_rows = self.sheet.max_row
        self.max_columns = self.sheet.max_column

    def getShift(self,place):

        appendToK = False
        appendToF = False

        shift ={}

        tam_kitchen=0
        tam_front=0

        kitchen_rows = []
        front_rows = []
        extracted = False
        for i in range(1,self.max_rows+1):
            cell_obj = self.sheet.cell(row=i,column=2)

            if cell_obj.value =="Kitchen":
                appendToK=True
                extracted= True
            elif cell_obj.value =="Front":
                appendToF=True
                appendToK = False
                extracted= True
            elif cell_obj.value == None:
                
                appendToF=False
                appendToK=False
                if extracted:
                    global endOfScheduleTable 
                    endOfScheduleTable = i
                    break

            
            if(appendToK):
                kitchen_rows.append(i)

            if(appendToF):
                front_rows.append(i)

        if place == "Kitchen":

            for row in kitchen_rows:
                cell_obj = self.sheet.cell(row=row,column=2) 
                if cell_obj.value=="Tam":
                    tam_kitchen = row

            for j in range(3, self.max_columns):
                cell_obj = self.sheet.cell(row=tam_kitchen, column=j)
                if cell_obj.value !=None:
                    shift[j-2]=cell_obj.value
        
        if place == "Front":

            for row in front_rows:
                cell_obj = self.sheet.cell(row=row,column=2) 
                if cell_obj.value=="Tam":
                    tam_front = row

            for j in range(3, self.max_columns):
                cell_obj = self.sheet.cell(row=tam_front, column=j)
                if cell_obj.value !=None:
                    shift[j-2]=cell_obj.value
        
        return shift

    def getLegend(self):

        kitchen_data = False
        front_data = False
        kitchen_legend = {}
        front_legend = {}

        for i in range(1,self.max_rows):
            cell_obj = self.sheet.cell(row=i,column=1)

            if(cell_obj.value=="Kitchen"):
                kitchen_data = True
                front_data = False
            
            if(kitchen_data):
                cell_obj = self.sheet.cell(row = i+1,column=1)

                if (cell_obj.value != "Front"):
                    keyValuePair = str(cell_obj.value).split(": ")
                    kitchen_legend[keyValuePair[0]]=keyValuePair[1]

            if(cell_obj.value=="Front"):
                kitchen_data = False
                front_data = True

            if( front_data):
                front_cell = self.sheet.cell(row = i+2,column=1)
                if(front_cell.value != None):
                    keyValuePair = str(front_cell.value).split(": ")
                    front_legend[keyValuePair[0]]=keyValuePair[1]
            
            if(cell_obj.value == None):
                kitchen_data = False
                front_data = False

        return kitchen_legend, front_legend

    def getSchedule(self,place):

        shift = self.getShift(place)

        kitchenLegend , frontLegent = self.getLegend()

        schedule = {}

        if place == "Kitchen":
            for date,tm in shift.items():
                if type(tm) == str:
                    schedule[date] = tm
                else:
                    schedule[date] = kitchenLegend[str(int(tm))]

        if place == "Front":
            for date,tm in shift.items():
                if type(tm) == str:
                    schedule[date] = tm
                else:
                    schedule[date] = frontLegent[str(int(tm))]

        return schedule

    def getDateTime(self,day,time):

        now = datetime.datetime.now()

        hoursMinutes = time.split(':')

        if(len(hoursMinutes) == 1):
            hoursMinutes.append("00")

        dt = datetime.datetime(now.year,now.month,day,int(hoursMinutes[0]),int(hoursMinutes[1]))

        return dt

    

    def getAuxLegend(self,col, letter):
        for i in range(endOfScheduleTable,self.max_rows+1):
            cell_obj = self.sheet.cell(row=i,column=col+2)
            keyVal = cell_obj.value.split(": ")
            if (keyVal[0]==letter):
                return keyVal[1]
